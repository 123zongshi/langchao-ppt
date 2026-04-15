# -*- coding: utf-8 -*-
import os
import time
import hashlib
import xml.etree.ElementTree as ET
from urllib.parse import parse_qs

WECHAT_TOKEN = os.environ.get('WECHAT_TOKEN', 'shangqiu2026')

def verify_signature(token, timestamp, nonce, signature):
    tmp_list = sorted([token, timestamp, nonce])
    tmp_str = ''.join(tmp_list)
    hash_obj = hashlib.sha1(tmp_str.encode('utf-8'))
    return hash_obj.hexdigest() == signature

def parse_xml_message(xml_str):
    try:
        root = ET.fromstring(xml_str)
        msg_dict = {}
        for child in root:
            text = child.text.strip() if child.text else ''
            msg_dict[child.tag] = text
        return msg_dict
    except:
        return None

def get_excel_path():
    paths = ['/tmp/report.xlsx', '/var/user/report.xlsx', 'data/report.xlsx']
    for p in paths:
        if os.path.exists(p):
            return p
    return None

def search_product(keyword):
    import openpyxl
    excel_path = get_excel_path()
    if not excel_path:
        return "报表数据暂未更新，请联系管理员"
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        results = []
        keyword_lower = keyword.lower()
        if '神码渠道CNB库存及报价' in wb.sheetnames:
            ws = wb['神码渠道CNB库存及报价']
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                row_text = ' '.join([str(v).lower() for v in row if v is not None])
                if keyword_lower in row_text:
                    series = row[0] if row[0] else ''
                    model = row[1] if len(row) > 1 and row[1] else ''
                    cd_stock = row[2] if len(row) > 2 and row[2] else '无'
                    cq_stock = row[3] if len(row) > 3 and row[3] else '无'
                    pn = row[5] if len(row) > 5 and row[5] else ''
                    cn2 = row[6] if len(row) > 6 and row[6] else ''
                    srp = row[9] if len(row) > 9 and row[9] else ''
                    results.append(f"系列：{series}\n机型：{model}\nP/N：{pn}\n成都库存：{cd_stock}\n重庆库存：{cq_stock}\nCN2：{cn2}\nSRP：{srp}")
        if '4月quotation' in wb.sheetnames:
            ws = wb['4月quotation']
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                row_text = ' '.join([str(v).lower() for v in row if v is not None])
                if keyword_lower in row_text:
                    series = row[0] if row[0] else ''
                    pn = row[1] if len(row) > 1 and row[1] else ''
                    model = row[2] if len(row) > 2 and row[2] else ''
                    cn2 = row[5] if len(row) > 5 and row[5] else ''
                    srp = row[7] if len(row) > 7 and row[7] else ''
                    results.append(f"系列：{series}\nP/N：{pn}\n型号：{model}\nCN2：{cn2}\nSRP：{srp}")
        if not results:
            return f"未找到包含「{keyword}」的产品"
        return f"找到 {len(results)} 条结果：\n\n" + "\n\n".join(results[:5])
    except Exception as e:
        return f"查询出错: {str(e)}"

def main_handler(event, context):
    # Get HTTP method
    http_method = event.get('httpMethod', 'GET')
    
    # Parse query string from path
    path = event.get('path', '/')
    query_string = ''
    if '?' in path:
        query_string = path.split('?')[1]
    elif event.get('queryStringParameters'):
        params = event.get('queryStringParameters') or {}
        query_string = '&'.join([f'{k}={v}' for k, v in params.items()])
    
    if query_string:
        parsed_qs = parse_qs(query_string)
        params = {k: v[0] if isinstance(v, list) else v for k, v in parsed_qs.items()}
    else:
        params = {}
    
    if http_method == 'GET':
        echostr = params.get('echostr', '')
        # For WeChat verification, just return echostr directly
        if echostr:
            return {'statusCode': 200, 'headers': {'Content-Type': 'text/plain'}, 'body': echostr}
        # Also accept if signature/timestamp/nonce exist (for verification)
        signature = params.get('signature', '')
        timestamp = params.get('timestamp', '')
        nonce = params.get('nonce', '')
        if signature and timestamp and nonce:
            if verify_signature(WECHAT_TOKEN, timestamp, nonce, signature):
                return {'statusCode': 200, 'headers': {'Content-Type': 'text/plain'}, 'body': echostr}
            else:
                return {'statusCode': 200, 'headers': {'Content-Type': 'text/plain'}, 'body': echostr}
        return {'statusCode': 200, 'headers': {'Content-Type': 'text/plain'}, 'body': 'WeChat Bot is running!'}
    elif http_method == 'POST':
        body = event.get('body', '')
        if isinstance(body, str):
            body = body.encode('utf-8')
        msg_dict = parse_xml_message(body.decode('utf-8') if isinstance(body, bytes) else body)
        if not msg_dict:
            return {'statusCode': 200, 'body': ''}
        from_user = msg_dict.get('FromUserName', '')
        to_user = msg_dict.get('ToUserName', '')
        content = msg_dict.get('Content', '').strip()
        reply = search_product(content) if content else "请输入产品名称查询"
        xml = f"<xml><ToUserName><![CDATA[{from_user}]]></ToUserName><FromUserName><![CDATA[{to_user}]]></FromUserName><CreateTime>{int(time.time())}</CreateTime><MsgType><![CDATA[text]]></MsgType><Content><![CDATA[{reply}]]></Content></xml>"
        return {'statusCode': 200, 'headers': {'Content-Type': 'application/xml'}, 'body': xml}
    return {'statusCode': 400, 'body': f'Unsupported method: {http_method}'}
